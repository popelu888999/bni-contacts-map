# -*- coding: utf-8 -*-
from openpyxl import load_workbook
import csv

wb = load_workbook(r'F:\work\BNI\BNI联系人清单.xlsx', data_only=True)
ws = wb.active

with open(r'F:\work\BNI\BNI联系人清单_temp.csv', 'w', encoding='utf-8-sig', newline='') as f:
    writer = csv.writer(f)
    for row in ws.iter_rows(values_only=True):
        writer.writerow(row)

print("Done!")
